from client import GraphQlClient
import luxor
import pandas as pd
import time
from openpyxl import load_workbook
import os

def load_params_from_txt(text_path):

    text_file = open(text_path)
    text = str(text_file.readline())
    text_file.close()

    return text

if __name__ == '__main__':

    # Declaring the main ant the params path
    user_name = os.getlogin()
    pre_user_path = 'C:/Users'
    post_user_path = 'Documents/Intiura_Miner_App'
    main_file_path = os.path.join(pre_user_path,user_name,post_user_path)
    params_file_path = os.path.join(pre_user_path,user_name,post_user_path, 'parametros')

    # Declaring other paths
    luxor_key_path = os.path.join(params_file_path,'luxor_key.txt')
    xlsx_path = os.path.join(main_file_path, 'Facturacion.xlsx')

    # -------------------------- Overrided ---------------------------
    xlsx_path = 'Facturacion.xlsx'

    # Parameters for the luxor API
    luxor_key = load_params_from_txt(luxor_key_path)
    method = "POST"
    host = "https://api.beta.luxor.tech/graphql"

    client  = GraphQlClient(
        host= host,
        method=method,
        key= luxor_key,
    )

    username = luxor.get_subaccounts(first=10)['data']['users']['edges'][0]['node']['username']

    # Values for the excel file
    summary = luxor.get_subaccount_mining_summary(username,'BTC','_1_HOUR')['data']['getMiningSummary']
    summary_values = [float(info) for info in list(summary.values())]
    summary_values[0] /= 10**15
    day = time.strftime("%Y/%m/%d,%H:%M:%S").split(',')[0]
    hour = time.strftime("%Y/%m/%d,%H:%M:%S").split(',')[1]
    summary_values = {'Day': [day],'Hour':[hour],'Hashrate':[summary_values[0]]}

    summary_values = pd.DataFrame(data = summary_values)

    # Writing on Excel the hourly data
    with pd.ExcelWriter(xlsx_path, mode = 'a', engine="openpyxl", if_sheet_exists = 'overlay') as writer:
        summary_values.to_excel(writer, sheet_name="data_diaria", startrow=writer.sheets['data_diaria'].max_row, header = False, index=False) 

    # If it is the las execution of the day 23:xx:xx write the daily average
    if hour[:2] == '23':

        data = pd.read_excel(xlsx_path, sheet_name='data_diaria')
        daily_data = data[data['Day'].str.contains(day)]
        average_hashrate = sum(daily_data['Hashrate'])/len(daily_data)

        daily_values = {'Day': [day],'Average Hashrate':[average_hashrate], 'USD': [average_hashrate*10]}
        daily_values = pd.DataFrame(data= daily_values)

        with pd.ExcelWriter(xlsx_path, mode = 'a', engine="openpyxl", if_sheet_exists = 'overlay') as writer:
            daily_values.to_excel(writer, sheet_name="data_mensual", startrow=writer.sheets['data_mensual'].max_row, header = False, index=False) 

        # check if it is the last day of the month
        day_timestamp = pd.Timestamp(int(day.split('/')[0]), int(day.split('/')[1]), int(day.split('/')[2]))

        if day_timestamp.is_month_end == True:

            month = day[:7]
            data = pd.read_excel(xlsx_path, sheet_name='data_mensual')
            monthly_data = data[data['Day'].str.contains(month)]
            total_usd = sum(monthly_data['USD'])

            monthly_values = {'month': [month], 'Total': [total_usd]}
            monthly_values = pd.DataFrame(data= monthly_values)

            with pd.ExcelWriter(xlsx_path, mode = 'a', engine="openpyxl", if_sheet_exists = 'overlay') as writer:
                monthly_values.to_excel(writer, sheet_name="factura", startrow=writer.sheets['factura'].max_row, header = False, index=False) 


    print('')
    # Open our existing CSV file in append mode
    # Create a file object for this file

    # DataFrame.to_excel(excel_writer, sheet_name='Sheet1', na_rep='', float_format=None, columns=None, header=True, index=True, index_label=None, startrow=0, startcol=0, engine=None, merge_cells=True, encoding=_NoDefault.no_default, inf_rep='inf', verbose=_NoDefault.no_default, freeze_panes=None, storage_options=None)
    
    
    # with open(csv_path, 'a',newline='') as f_object:
    
    #     # Pass this file object to csv.writer()
    #     # and get a writer object
    #     writer_object = writer(f_object)
    
    #     # Pass the list as an argument into
    #     # the writerow()
    #     writer_object.writerow(summary_values)
    
    #     # Close the file object
    #     f_object.close()

    # subaccount: ikerinti,
    # mpn: BTC,
    # input_interval: _15_MINUTE, _1_HOUR, _1_HOUR and _1_DAY